from ossapi import Ossapi
from openpyxl import load_workbook
import commons

def read_data():
    wb = load_workbook(filename='sheets/qualifier.xlsx', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    if m_row == None:
        # we will defined an default max mappools size if we couldn't
        # get max_row when reading , usually by xlsx files saved by 3rd
        # software like "google sheet"
        m_row = 1000
    participates = []
    config = {
        'is_teammode':ws.cell(row=2, column=5).value,
        'sum_cond_index':ws.cell(row=1, column=9).value,
        'avg_cond_index': ws.cell(row=2, column=9).value
    }

    cond_headers = []
    map_headers = []
    for i in range (0,5):
        cond_header = ws.cell(row=5, column=6 + i).value
        if cond_header is None:
            if i == config['sum_cond_index'] or i == config['avg_cond_index']:
                continue
            else:
                break
        cond_headers.append(cond_header)

    for i in range (0,16):
        map_header = ws.cell(row=5, column=12 + i * 2).value
        if map_header is None:
            break
        map_headers.append(map_header)

    for i in range(8, m_row + 1):
        participate = {
            'place': ws.cell(row=i, column=2).value,
            'id': ws.cell(row=i, column=3).value,
            'name': ws.cell(row=i, column=4).value,
            'link': ws.cell(row=i, column=5).value,
            'bg': ws.cell(row=i, column=11).value,
            'conds':[],
            'maps':[]
        }
        if participate['name'] is None:
            break
        if not isinstance(participate['name'],str):
            participate['name'] = str(participate['name'])

        for j in range(0,len(cond_headers)):
            condtion = ws.cell(row=i, column=6+j).value
            if not (condtion is None):
                participate['conds'].append(condtion)

        for j in range(0,len(map_headers)):
            map_score = ws.cell(row=i, column=12 + j*2 + 1).value
            map_place = ws.cell(row=i, column=12 + j*2).value
            if map_score is None:
                break
            else:
                if map_place is None:
                    map_place = ''
                map = {
                    'score': map_score,
                    'place': map_place
                }
                participate['maps'].append(map)
        participates.append(participate)
    return config, cond_headers, map_headers, participates

def draw_qualifier_table_header(config, cond_headers, map_headers):
    result = '{{QualifierTable\n'
    result_configs = ''
    is_config_exist = False
    if not config['is_teammode']:
        result_configs += '|playerMode=1 '
        is_config_exist = is_config_exist or True
    if not config['sum_cond_index'] is None:
        result_configs += '|sumCond='+ str(config['sum_cond_index'])+' '
        is_config_exist = is_config_exist or True
    if not config['avg_cond_index'] is None:
        result_configs += '|avgCond=' + str(config['avg_cond_index']) + ' '
        is_config_exist = is_config_exist or True
    if is_config_exist:
        result_configs += '\n'

    result_conds = ''
    for index,cond_header in enumerate(cond_headers):
        result_conds += '|cond' + str(index+1) + '=' + cond_header + ' '
    if len(cond_headers) > 0:
        result_conds += '\n'

    result_maps = ''
    for index,map_header in enumerate(map_headers):
        result_maps += '|map' + str(index+1) + '=' + map_header + ' '
    if len(cond_headers) > 0:
        result_maps += '\n'
    return result + result_configs + result_conds + result_maps


def draw_qualifier_participate_team(participate):
    result = '|{{Json|team=' + participate['name']
    if not participate['link'] is None:
        result += '|link=' + participate['link']
    if not participate['place'] is None:
        result += '|place=' + str(participate['place'])
    if not participate['bg'] is None:
        result += '|bg=' + participate['bg']
    for index, condRating in enumerate(participate['conds']):
        result += '|cond' + str(index+1) + '=' + f'{condRating:,}'
    for index, map in enumerate(participate['maps']):
        map_score = map['score']
        result += '|map' + str(index+1) + '=' + str(map['place']) + ';' + f'{map_score:,}'
    result += '}}\n'
    return result

def draw_qualifier_participates_solo(participate):
    player_id = participate['name']
    player_flag = commons.get_player_osuflag(player_id)
    player_id_display = commons.clean_clan_tags(player_id)
    result = '|{{Json|player=' + player_id_display
    # TODO: Remove mode: Remove all '[/]'
    if player_id_display.startswith('[') and (participate['link'] is None):
        result += '|link=' + player_id_display.replace("[", "(").replace("]", ")")
    if not player_flag is None:
        result += '|flag=' + player_flag
    if not participate['link'] is None:
        result += '|link=' + participate['link']
    if not participate['place'] is None:
        result += '|place=' + str(participate['place'])
    if not participate['bg'] is None:
        result += '|bg=' + participate['bg']
    for index, condRating in enumerate(participate['conds']):
        result += '|cond' + str(index+1) + '=' + f'{condRating:,}'
    for index, map in enumerate(participate['maps']):
        map_score = map['score']
        result += '|map' + str(index + 1) + '=' + str(map['place']) + ';' + f'{map_score:,}'
    result += '}}\n'
    return result

def draw_qualifier_table_tail():
    return '}}\n'

if __name__ == '__main__':
    config, cond_headers, map_headers, participates = read_data()
    participates_count = len(participates)
    with open('reuslt_qualifier.txt','w',encoding='UTF-8') as resultFile:
        header = draw_qualifier_table_header(config, cond_headers, map_headers)
        tail = draw_qualifier_table_tail()
        resultFile.writelines(header + '\n')
        if config['is_teammode']:
            for participate in participates:
                resultFile.writelines(draw_qualifier_participate_team(participate))
        else:
            for participate in participates:
                resultFile.writelines(draw_qualifier_participates_solo(participate))
        resultFile.writelines('\n')
        resultFile.writelines(tail)