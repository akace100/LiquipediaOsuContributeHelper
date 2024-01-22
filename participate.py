import ossapi
from ossapi import Ossapi
from openpyxl import load_workbook
from openpyxl import Workbook
import commons

api = commons.generate_osu_api()

def read_participates():
    wb = load_workbook(filename='sheets/participate.xlsx', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    if m_row == None:
        # we will defined an default max mappools size if we couldn't
        # get max_row when reading , usually by xlsx files saved by 3rd
        # software like "google sheet"
        m_row = 400
    participates = []
    # Solo participate modes on if no team name.
    is_solo_participate = ws.cell(row=2, column=1).value is None
    for i in range(2, m_row + 1):
        participate = {}
        if not is_solo_participate:
            participate['name'] = ws.cell(row=i, column=1).value
        participate['players'] = ws.cell(row=i, column=2).value
        if participate['players'] is None:
            break
        if isinstance(participate['players'],int):
            participate['players'] = str(participate['players'])
        participate['qualifier'] = ws.cell(row=i, column=3).value
        participate['link'] = ws.cell(row=i, column=4).value
        participate['image'] = ws.cell(row=i, column=5).value
        participates.append(participate)
    return participates, is_solo_participate


def generateTeamCardInfo(team):
    result = '{{TeamCard\n' + \
             '|team=' + team['name'] + '\n'
    if not team['link'] is None:
        result += '|link=' + team['link'] + '\n'
    if not team['image'] is None:
        result += '|image=' + team['image'] + '\n'
    players = commons.clean_string(team['players']).split(',')
    for i in range(0,len(players)):
        player = str(players[i])
        # remove possible head space
        while player.startswith(' '):
            player = player[1:]
        print('generate ' + player + '\'s info...')
        try:
            player_Country = api.user(player).country_code
        except Exception as e:
            print('error finding player country:' + player)
            print(e)
            player_Country = ''
        player_name_clean = commons.clean_clan_tags(player)
        result +=f'|p{i+1}={player_name_clean} |p{i+1}flag={player_Country}'
        if '[' in player_name_clean and ']' in player_name_clean:
            parsed_player_name = player_name_clean.replace("[","(").replace("]",")")
            result +=f' |p{i+1}link={parsed_player_name}'
        if i == 0:
            result += ' |captain=true\n'
        else:
            result += '\n'
    if not team['qualifier'] is None:
        result += '|qualifier=' + team['qualifier'] + '\n'
    result += '}}\n'
    return result

def generateSoloInfo(player_id):
    print('generate ' + player_id + 's info...')
    result = '     |{{1Opponent|'
    player_id_clean = commons.clean_clan_tags(player_id)
    flag = commons.get_player_osuflag(player_id)
    result += player_id_clean + '|flag=' + flag + '}}'
    if '[' in player_id_clean and ']' in player_id_clean:
        parsed_player_id = player_id_clean.replace("[", "(").replace("]", ")")
        result += f'|link={parsed_player_id}'
    return result


if __name__ == '__main__':
    participates, is_solo_participate = read_participates()
    if is_solo_participate:
        players = participates
        result = ''
        for i in range(0,len(players)):
            player_id = players[i]['players']
            result += generateSoloInfo(player_id)
            if i < len(players):
                result += '\n'

    else:
        teams = participates
        result = ''
        for i in range(0, len(teams)):
            # print('generate ' + teams[i]['name'] + 's info...')
            result += generateTeamCardInfo(teams[i])
            if i < len(teams):
                result += '\n'

    f = open("result_participate.txt", "w", encoding='utf-8')
    f.write(result)
    f.close()

