from openpyxl import load_workbook
import commons

api = commons.generate_osu_api()
default_max_participates_in_list = 9

def read_staffs():
    staffs = {}
    wb = load_workbook(filename='sheets/staff.xlsx', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    max_participates_in_list = ws['J3']
    if max_participates_in_list is None:
        max_participates_in_list = default_max_participates_in_list
    for i in range(3, m_row + 1):
        position = ws.cell(row=i, column=2).value
        if position is None:
            # incase possible staff list shorter that Note
            break
        participates = ws.cell(row=i, column=3).value.split(',')
        if position not in staffs.keys():
            staffs[position] = []
        for participate in participates:
            # remove possible duplicated space (not in id's ,might start with?)
            # participate = participate.replace(" ","")
            staffs[position].append(participate)
    return staffs

def generate_broadcaster_card(position,staffs):
    result = '{{BroadcasterCard\n|position='+position+'\n'
    for i in range(0, len(staffs)):
        staff_id = staffs[i]
        staff_flag = commons.get_player_osuflag(staff_id)
        result += f'|b{i}={commons.clean_clan_tags(staff_id)}|b{i}flag={staff_flag}\n'
    result += '}}\n'
    return result

# Tabs will be divided automaticlly:
def set_belong_tabs(staffs):
    talent_positions = ['Streamer','Commentator','Desk Host','Analyst']
    tabs = {}
    for position in staffs.keys():
        for talent_position in talent_positions:
            if talent_position in position:
                print ('cools')
    return tabs.keys()


# def generate_broadcast_talent_tabhead(positions):
#     tabs = getTabs(positions)
#     if tabs is None:
#         return None
#     result = '{{Tabs dynamic\n'
#     for i in range(0, positions):
#         result += f

if __name__ == '__main__':
    staffs = read_staffs()
    positions = staffs.keys()
    # set_belong_tabs(staffs)
    for position in positions:
        print(generate_broadcaster_card(position,staffs[position]))
        print('\n')
    # for position in positions:
    #

