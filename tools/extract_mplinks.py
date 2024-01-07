from datetime import datetime
import time

import ossapi
import os

import pytz as pytz
from ossapi import Ossapi, MatchEventType
from openpyxl import load_workbook
from openpyxl import Workbook
import commons

api = commons.generate_osu_api()


def read_mplinks():
    mplinks = []
    wb = load_workbook(filename='..\\sheets\\extract_urls.xlsx', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    mplink_column_index = 2
    if m_row is None:
        m_row = 150
    for i in range(3, m_row + 1):
        mplink = ws.cell(row=i, column=mplink_column_index).value
        if mplink is None:
            break
        else:
            mplinks.append(mplink)
    return mplinks


def read_match(mplink):
    try:
        match = api.match(mplink)
    except ValueError as e:
        print('Invalid mplink : ' + str(mplink))
        return None

    # get full events
    current_events = match.events
    events = current_events
    first_event_ID = events[0].id
    while (len(current_events) == 100):
        current_events = api.match(mplink, before=first_event_ID).events
        if len(current_events) == 0:
            break
        else:
            first_event_ID = current_events[0].id
            new_events = list(current_events)
            new_events.extend(events)
            events = new_events
    match.events = events

    return match


def get_link_id(mplink):
    link_id = mplink.split('/')[-1]
    return link_id


def get_games(match):
    games = []
    for event in match.events:
        if event.detail.type == MatchEventType.OTHER:
            game = event.game
            games.append(game)
    return games


def get_match_creator(match):
    create_event = match.events[0]
    if create_event.user_id is None:
        return 'unknown'
    else:
        creator_username = api.user(create_event.user_id).username
    return creator_username

def get_about_time(target_time):
    timestamp = datetime.timestamp(target_time)
    timestamp_about = timestamp // (15 * 60) * (15 * 60)
    if timestamp % (15 * 60) > 7.5 * 60:
        timestamp_about = timestamp_about + (15 * 60)
    return datetime.fromtimestamp(timestamp_about,tz=pytz.timezone("UTC"))

def print_extract_dates(match_infos):
    with open('..\\result_extract_urls.txt', 'w') as file:
        for match_info in match_infos:
            file.writelines(f'mplink : ' + match_info['mplink'] + '\n')
            if match_info['roomname'] is None:
                file.writelines(f'Invalid mplink\n')
            else:
                file.writelines('roomname : ' + match_info['roomname'] + '\n')
                file.writelines('start_time_excat : ' + match_info['start_time_exact'] + '\n')
                file.writelines('start_time_about : ' + match_info['start_time_about'] + '\n')
                file.writelines('referee : ' +
                                (match_info['referee'] if match_info['referee'] is not None else 'unknown') + '\n')
            file.writelines('\n')
    with open('..\\result_extract_urls.csv', 'w') as file:
        file.writelines('mplink,roomname,start_time_excat,start_time_about,referee')
        for match_info in match_infos:
            file.writelines(f'mplink : ' + match_info['mplink'] + ',')
            if match_info['roomname'] is None:
                file.writelines(f'Invalid mplink')
            else:
                file.writelines(match_info['roomname'] + ',')
                file.writelines(match_info['start_time_exact'] + ',')
                file.writelines(match_info['start_time_about'] + ',')
                file.writelines((match_info['referee'] if match_info['referee'] is not None else 'unknown'))
            file.writelines('\n')



if __name__ == '__main__':
    mplinks = read_mplinks()
    match_infos = []
    for mplink in mplinks:
        match_info = {
            'mplink': mplink,
            'roomname': None,
            'start_time_exact': None,
            'start_time_about': None,
            'referee': None
        }
        link_id = get_link_id(mplink)
        match = read_match(get_link_id(mplink))
        if match is not None:
            games = get_games(match)
            first_game_time = games[0].start_time
            match_info['roomname'] = match.match.name
            match_info['start_time_exact'] = first_game_time.strftime("%Y-%m-%d %H:%M:%S  {{Abbr/UTC}}")
            match_info['start_time_about'] = get_about_time(first_game_time).strftime("%Y-%m-%d %H:%M:%S  {{Abbr/UTC}}")
            match_info['referee'] = get_match_creator(match)
        match_infos.append(match_info)
    print_extract_dates(match_infos)
