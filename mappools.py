from ossapi import Ossapi
from openpyxl import load_workbook
import commons

api = commons.generate_osu_api()

def readMapools():
    Mappools = {}
    wb = load_workbook(filename='sheets/mappools.xlsx', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    if m_row == None:
        # we will defined an default max mappools size if we couldn't
        # get max_row when reading , usually by xlsx files saved by 3rd
        # software like "google sheet"
        m_row = 50
    for i in range(2, m_row + 1):
        Mod = ws.cell(row=i, column=1).value
        ID = ws.cell(row=i, column=2).value
        BID = ws.cell(row=i, column=3).value
        if ID is None :
            break
        if Mod is not None:
            if not Mod in Mappools.keys():
                Mappools[Mod] = []
            Mappools[Mod].append({'id': ID, 'bid': BID})
        else:
            # For cases Mappools isn't split by Mods (early World Cup, Qualifier Stage)
            if not '' in Mappools.keys():
                Mappools[''] = []
            Mappools[''].append({'id': ID, 'bid': BID})
    return Mappools

def generate_MappoolsHeader(mappools):
    head = '{{Tabs dynamic\n'
    tail = '|This=1\n}}\n'
    result = head
    mods = list(mappools.keys())
    for i in range(0,len(mods)):
        mod = mods[i]
        result += f'|name{i+1}={mod}\n'
    return result+tail

def generate_Mappools(mappools, headless = False):
    mods = list(mappools.keys())
    result = ''
    for i in range(0,len(mods)):
        mod = mods[i]
        beatmapIDs = mappools[mod]
        if not headless:
            head = '{{Tabs dynamic/tab|' + str(i + 1) + '}}\n{{box|start|padding=1em}}\n'
            tail = '{{box|end}}\n'
        else:
            head = ''
            tail = ''
        result += head
        for beatmapID in beatmapIDs:
            id = beatmapID['id']
            bid = beatmapID['bid']
            url, info = get_beatmapInfo(bid)
            mapinfo = '* \'\'\''+id+'\'\'\' : '
            if id.startswith('TB'):
                mapinfo += '\'\'\'['+url+' '+info+']\'\'\'\n'
            else:
                mapinfo += '[' + url + ' ' + info + ']\n'
            result += mapinfo
        result += tail
    return result

def get_beatmapInfo(betmapid):
    try:
        beatmap = api.beatmap(betmapid)
        beatmapset = beatmap.beatmapset()
        cleantags = ['[4K] ', '[7K] ']
        beatmap_version = beatmap.version
        for cleantag in cleantags:
            beatmap_version = beatmap_version.replace(cleantag, '')
        url = beatmap.url
        info = f'<nowiki>{beatmapset.artist} - {beatmapset.title} ({beatmapset.creator}) [{beatmap_version}] </nowiki>'
    except ValueError as e:
        print(e)
        url=''
        info=''
    return url, info

if __name__ == '__main__':
    mappools = readMapools()
    mods = list(mappools.keys())
    if mods[0] == '' and len(mods) == 1:
        head = ''
        body = generate_Mappools(mappools, headless=True)
        tail = ''
    else:
        head = generate_MappoolsHeader(mappools)
        body = generate_Mappools(mappools)
        tail = '{{Tabs dynamic/end}}\n'

    f = open("result_mappool.txt", "w", encoding='utf-8')
    f.write(head + body + tail)
    f.close()

