import re

import requests
from ossapi import Mod
from ossapi import Ossapi
from ossapi import MatchEventType
from ossapi import TeamType
from openpyxl import load_workbook
import commons

client_id = commons.client_id
client_secret = commons.client_secret
api = commons.generate_osu_api()

def getMatch_native(mpID):
    url = "https://osu.ppy.sh/oauth/token"
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    body = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": 'client_credentials',
        "scope": 'public'
    }
    response = requests.post(url, headers=headers, data=body)
    access_token = response.json()['access_token']

    url = f"https://osu.ppy.sh/api/v2/matches/{mpID}"
    headers = {
        'Authorization': 'Bearer ' + access_token,
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    return response.json()

def readDatas():
    mappools = []
    modMultipliers = []
    players = []
    mplinks = []
    settings = {}
    wb = load_workbook(filename='sheets/match_result.xlsm', read_only=True)
    ws = wb.active
    m_row = ws.max_row
    if m_row == None:
        # we will defined an default max mappools size if we couldn't
        # get max_row when reading , usually by xlsx files saved by 3rd
        # software like "google sheet"
        m_row = 150
    columnIndex = {
        'mappool_id' : 2,
        'mappool_bid' : 3,
        'modMultiplier_mod' : 5,
        'modMultiplier_multiplier' : 6,
        'playerUID' : 8,
        'mplink_columnIndex' : 10
    }
    finshedRead = {
        'mappool' : False,
        'modMultiplier' : False,
        'playerList' : False,
        'mplink' : False
    }
    for i in range(4, m_row + 1):
        readMark = 0
        if not finshedRead['mappool']:
            readMark += 1
            map = {
                'id' : ws.cell(row=i, column=columnIndex['mappool_id']).value,
                'bid' : ws.cell(row=i, column=columnIndex['mappool_bid']).value
            }
            if (map['id'] is None or map['bid'] is None):
                finshedRead['mappool'] = True
            else:
                mappools.append(map)
        if not finshedRead['modMultiplier']:
            readMark += 1
            multiplier = {
                'mod' : ws.cell(row=i, column=columnIndex['modMultiplier_mod']).value,
                'multiplier' : ws.cell(row=i, column=columnIndex['modMultiplier_multiplier']).value
            }
            if (multiplier['mod'] is None or multiplier['multiplier'] is None):
                finshedRead['modMultiplier'] = True
            else:
                modMultipliers.append(multiplier)
        if not finshedRead['playerList']:
            readMark += 1
            playerUID = ws.cell(row=i, column=columnIndex['playerUID']).value
            if playerUID is None:
                finshedRead['playerList'] = True
            else:
                players.append(playerUID)
        if not finshedRead['mplink']:
            readMark += 1
            mplink = ws.cell(row=i, column=columnIndex['mplink_columnIndex']).value
            if mplink is None:
                finshedRead['mplink'] = True
            else:
                mplinks.append(mplink)
        # stop reading unused information
        if (readMark == 0):
            break
    settings['swap_teams'] = ws['O4'].value
    settings['win_conditions'] = ws['O7'].value
    return mappools,modMultipliers,players,mplinks,settings


#RoomName Following Format: {Match Abbr}:({TeamA}) vs ({TeamsB})
#Not considering any player / teams has ) vs (
def resolveTeamNames(roomName):
    # we only need first one ':(' to take the first team/player's header
    # cuz considing teams also have that.
    colonIndex = roomName.find(': (')
    teamNameSplits = roomName[colonIndex:].split(') vs (')
    if len (teamNameSplits) < 2:
        raise ValueError('Room name isn\'t format')
    if len(teamNameSplits) > 2:
        raise ValueError('Player / Team name contains ") vs (", cause split problem')
    return teamNameSplits[0][3:],teamNameSplits[1][:-1]

def findmapIDbyBID(bid,modlist):
    for mod in modlist:
        if (mod['bid'] == bid):
            return mod['id']
    return None

def getPlayCount(game):
    if game.team_type == TeamType.HEAD_TO_HEAD:
        return 1
    if game.team_type == TeamType.TEAM_VS:
        countRed = 0
        countBlue = 0
        for score in game.scores:
            # exclude in-game referee
            if score.score > 0:
                if score.match.team == 'red':
                    countRed += 1
                if score.match.team == 'blue':
                    countBlue += 1
        if countRed != countBlue:
            #inconsist players means match invalid for count match size.
            return None
        else:
            return countRed

# input:
def get_player_slot(red_player_id,blue_player_id):
    return 0,1

def get_player_userid(room_name):
    red_player_userid = 0
    blue_player_userid = 0
    usernames = re.search(r'' , room_name)
    print(usernames.group(1),usernames.group(2))
    return red_player_userid,blue_player_userid

# input:
# possible ruleset = {{Mod} NF = x1.0} etc.
# return {redScore= blueScore=}
def getScore(game,modMultipliers,determinedByID = False,accuracyScore = False):
    redScore = None
    blueScore = None
    # For 1v1 tournaments
    if game.team_type == TeamType.HEAD_TO_HEAD:
        for score in game.scores:
            # common way : red player = slot 0; blue player = slot 1
            if determinedByID:
                if score.user_id == game.red_player_userid:
                    redScore = score.score if not accuracyScore else (score.accuracy * 100)
                if score.user_id == game.blue_player_userid:
                    blueScore = score.score if not accuracyScore else (score.accuracy * 100)
            else:
                if score.match.slot == 0:
                    redScore = score.score if not accuracyScore else (score.accuracy * 100)
                if score.match.slot == 1:
                    blueScore = score.score if not accuracyScore else (score.accuracy * 100)
    # for teamVS tournaments (include 1v1 teams)
    if (game.team_type == TeamType.TEAM_VS) or (game.team_type == TeamType.TAG_TEAM_VS):
        redScore = 0
        blueScore = 0
        redPlayers = 0
        bluePlayers = 0
        for score in game.scores:
            finalScore = score.score if not accuracyScore else score.accuracy
            for modMultiplier in modMultipliers:
                if Mod(modMultiplier['mod']) in score.mods:
                    finalScore = finalScore * modMultiplier['multiplier']
            # assume no invalid players in rooms
            if score.match.team == 'red':
                redScore += finalScore
                redPlayers += 1
            if score.match.team == 'blue':
                blueScore += finalScore
                bluePlayers += 1
        if accuracyScore:
            redScore = (redScore * 100) / redPlayers
            blueScore = (blueScore * 100) / bluePlayers
    return redScore, blueScore

def getFullEvents(match):
    match = api.match(mplink)
    current_events = match.events
    events = current_events
    first_event_ID = events[0].id
    while (len(current_events) == 100):
        current_events = api.match(mplink,before_id=first_event_ID).events
        if len(current_events) == 0:
            break
        else:
            first_event_ID = current_events[0].id
            new_events = list(current_events)
            new_events.extend(events)
            events = new_events
    return events

def getGames(match):
    games = []
    for event in match.events:
        if event.detail.type == MatchEventType.OTHER:
            game = event.game
            games.append(game)
    return games

if __name__ == '__main__':
    mappools,modMultipliers,players,mplinks,settings = readDatas()
    resultFile = open("result_match.txt", 'w',encoding='UTF-8')
    for mplink in mplinks:
        try:
            match = api.match(mplink)
        except ValueError as e:
            print('Invalid mplink : ' + str(mplink))
            resultFile.writelines(f'Invalid mplink {mplink}\n')
            resultFile.writelines('\n')
            continue
        events = getFullEvents(match)
        match.events = events
        print(f'loading https://osu.ppy.sh/community/matches/{mplink}')
        roomName = match.match.name
        # red_player_userid, blue_player_userid = get_player_userid(match.match.name)
        games = getGames(match)
        playerCount = None
        lastMapID = ''
        last_teamwin = ''
        mapresults = []
        red_games_win, blue_games_win = 0,0
        accuracyWin = settings['win_conditions'] == "ACC"
        for game in games:
            # TODO: consider old matches that maps have been removed
            mapID = findmapIDbyBID(game.beatmap_id,mappools)
            # ignore warmup maps
            if (mapID is None):
                continue
            if (playerCount is None):
                # count player each side to determined teamsize, so we could exclude "TB for fun".
                # not 100% for some same size TB or 1v1 "TB for fun".
                playerCount = getPlayCount(game)
            else:
                if ('TB' in mapID):
                    # exclude "TB for fun".
                    if (red_games_win != blue_games_win):
                        continue
            redScore , blueScore = getScore(game,modMultipliers,accuracyScore=accuracyWin)
            # exclude broken match
            if (redScore is None or blueScore is None):
                print(f'broken match detected in match : {roomName}')
            else:
                if settings['swap_teams']:
                    blueScore , redScore = redScore, blueScore
                mapresult = {
                    'map': mapID,
                    'mode': '',  # leave for future liquipedia updates
                    'score1': f'{redScore:,}' if not accuracyWin else "%.2f" % redScore + "%",
                    'score2': f'{blueScore:,}'if not accuracyWin else "%.2f" % blueScore + "%",
                    'winner': '1' if (redScore > blueScore) else '2'
                }
                # for rematch happen, only record last one, so the former result will be revent
                if lastMapID == mapID:
                    mapresults[-1] = mapresult
                    if last_teamwin == 'red':
                        red_games_win -= 1
                    if last_teamwin == 'blue':
                        blue_games_win -= 1
                else:
                    mapresults.append(mapresult)
                    lastMapID = mapID
                if (redScore > blueScore):
                    red_games_win += 1
                    last_teamwin = 'red'
                else:
                    blue_games_win += 1
                    last_teamwin = 'blue'
        # write files
        resultFile.writelines(f'========={roomName}========== {mplink}\n')
        for i in range(0,len(mapresults)):
            mapresult = mapresults[i]
            resultFile.writelines('    |map'+str(i+1)+'={{Map|map='+mapresult['map']+'' \
                                  '|mode='+mapresult['mode']+'|score1='+mapresult['score1']+''\
                                  '|score2='+mapresult['score2']+'|winner='+mapresult['winner']+'}}\n')
        resultFile.writelines('\n')
    resultFile.close()


